from openpyxl import *  # Workbook
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from HtmlHack.settings import MEDIA_ROOT
from openpyxl.utils import get_column_letter
from collections import defaultdict

class Excell:
    # openpyxl
    nombre = ''

    def __init__(self, nombre_):
        self.lista = []
        global nombre
        nombre = nombre_
        print('NombreFichero: ' + nombre )
        try:
            self.wb = openpyxl.load_workbook(nombre)  # workbook()
            hoja = self.mostrar_sheets()  #imprime hojas
            self.ws = self.wb[ hoja[0] ]

            #print('se abrio fichero excell '+nombre +' correctamente.')
        except:
            self.wb = openpyxl.Workbook()  #cambiado de workbook()
            #self.ws = self.wb.create_sheet()
            hoja = self.mostrar_sheets()  #imprime hojas
            
            self.ws = self.wb[ hoja[0] ]
            self.ws.title = hoja[0]
            
            self.salvarexcell2() #self.salvarexcell2()
            #print('se ha creado el fichero excell '+nombre+' correctamente.')
        # global ws
        self.ws = self.wb.active

    def createsheet(self, titulo):

        hoja = self.wb.create_sheet()
        hoja.title = titulo
        # hoja.sheet_properties.tabcolor = "aaaaaa"
        #self.wb.save(nombre)

    def deleteSheet(self, nombre):
        #hoja= self.wb.get_sheet_by_name(nombre)
        del self.wb[nombre]

    def getNombreColumna( nCol): #consigue la letra
        return get_column_letter(nCol)

    def cambiar_hoja(self, nombre):
        self.ws = self.wb[nombre]

    def mostrar_sheets(self):
        # print( wb.get_sheet_names( ) )
        h = self.wb.sheetnames
        print(h)
        return h

    def mostrar_celda(self, fila, columna):
        return self.ws.cell(row=fila, column=columna).value

    def ver_celda(self, celda):
        return self.ws[celda].value

    def escribir_celda(self, celda, val):
        # self.ws = self.wb[ 'tr' ]     #get_sheet_by_name("prb1")
        self.ws[celda].value = val
        #self.wb.save(nombre)

    def leer_rango(self, rfila, rcolumna):
        ret = []
        lista = self.ws[rfila: rcolumna]
        for i in lista:
            for j in i:
                ret.append(j.value)
        return ret

    def leer_fichero(self):
        ret = []
        item = []
        lista = self.ws.rows
        for i in lista:
            for j in i:
                item.append( j.value )
            ret.append( item )
            item = None
            item = []
        ret.pop( 0 )
        return ret

    def leer_fichero2(self):
        #lista = self.ws.rows
        #return lista
        cmps = None
        self.lista = []
        mtx = self.ws.rows
        try:
            for row in mtx:
                if row != '':
                    m = row[0].value
                    e = row[1].value
                    u = row[2].value
                    d = row[3].value
                    codigo = row[4].value
                    nombre = row[5].value
                    pacto = row[6].value
                    minimo = row[7].value
                    dc = row[8].value
                    gfh = row[9].value
                    disp = row[10].value
                    hosp = row[11].value
                    cmps = campos(m,e,u,d,codigo,nombre,pacto,minimo,dc,gfh,disp,hosp)
                    #cmps = (m+'#'+e+'#'+u+'#'+d+'#'+codigo+'#'+nombre+'#'+pacto+'#'+minimo+'#'+dc+'#'+gfh+'#'+disp+'#'+hosp).split('#')
                    self.lista.append(cmps)
        except Exception as e:
            return []

        return self.lista[1 : ]


    def leer_fila(self):
        ret = []
        lista = self.ws.rows
        for i in lista:
            for j in i:
                ret.append(j.value)
        return ret

    def leer_columna(self):
        ret = []
        lista = self.ws.columns
        for i in lista:
            for j in i:
                ret.append(j.value)
        return ret

    def merge(self, rfila, rcolumna):
        rango = rfila + ':' + rcolumna
        self.ws.merge_cells(rango)
        

    def unmerge(self, rfila, rcolumna):
        rango = rfila + ':' + rcolumna
        self.ws.unmerge_cells(rango)
        

    def insertar_imagen(self, ruta, celda):
        img = openpyxl.drawing.image.image(ruta)
        self.ws.add_image(img, celda)
        

    def insertar_rangofila(self, rango, fila, columna):
        for i, value in enumerate(rango):
            self.ws.cell(column=columna + i, row=fila, value=value)
            #print('Fila: '+ str(fila) + '  Columna: '+ str(columna) + '  Value: '+ str(value))

    def insertar_rangocolumna(self, rango, fila, columna):
        for i, value in enumerate(rango):
            self.ws.cell(column=columna, row=fila + i, value=value)
        

    def cambiarcolorfila(self, fila, col, inicol, fincol):
        for i in range(col):
            self.ws.cell(row=fila, column=i + 1).fill = patternfill(start_color=inicol, end_color=fincol,
                                                                    fill_type='solid')
        

    def cambiarcolorcolumna(self, fila, col, inicol, fincol):
        for i in range(fila):
            self.ws.cell(row=i + 1, column=col).fill = patternfill(start_color=inicol, end_color=fincol, fill_type='solid')
        

    def getnumerofilas(self):
        return self.ws.max_row

    def getnumerocolumnas(self):
        return self.ws.max_column

    def salvarexcell(self):
        self.wb.save( nombre )

    def salvarexcell2(self):
        self.wb.save( MEDIA_ROOT +'/' + nombre +'.xlsx' )
        
    def salvarexcell3(self):
        self.wb.save( MEDIA_ROOT +'/' + nombre +'.xlsx' )

    def __repr__(self):
        return str(self.lista)

    def __str__(self):
        return str(self.lista)

    
class campos:
    
    def __init__(self,mod, est, ubc, div, cod, nom, pac, mim , dc, gfh, disp, hosp ):
        self.modulo = mod
        self.estanteria = est
        self.ubicacion = ubc
        self.division = div
        self.codigo = cod
        self.nombre = nom
        self.pacto = pac
        self.minimo = mim
        self.dc = dc
        self.gfh = gfh
        self.dispositivo = disp
        self.hospital = hosp

#region revision
class ListaToQuerySet(list):

    def __init__(self, *args, model, **kwargs):
        self.model = model
        super().__init__(*args, **kwargs)

    def filter(self, *args, **kwargs):
        return self  # filter ignoring, but you can impl custom filter

    def order_by(self, *args, **kwargs):
        return self

def list_to_queryset(model, data):
    from django.db.models.base import ModelBase

    if not isinstance(model, ModelBase):
        raise ValueError("%s must be Model" % model)

    if not isinstance(data, list):
        raise ValueError("%s must be List Object" % data)

    pk_list = [obj.pk for obj in data]
    return model.objects.filter(pk__in=pk_list)

#endregion

class comprobarExcel:
    ListaUb = []
    ListaCo = []
    Lista = []
    ubicacion = ''
    codigo = ''

    def __init__(self, lista ):
        #print(lista)
        self.ListaUb.clear()
        self.Lista = lista
        for i in lista:
            ubicacion = str(i.modulo)+'-'+str(i.estanteria)+'-'+str(i.ubicacion)+'-'+str(i.division)
            codigo = str(i.codigo) + ' ' + str(i.dc)  # codigo + DC
            self.ListaUb.append( ubicacion )
            self.ListaCo.append( codigo )
            ubicacion = None
            codigo = None
        

    def comprobarDuplicados( self ):
        duplic = []
        ind = 0
        ind2 = 1
        #print('ListaUB: ' + str( len(self.ListaUb)))
        for i in self.ListaUb:
            ind2 = ind + 1
            for j in range(  len( self.ListaUb ) - ind2 ):
                if ind2 >= len( self.ListaUb ):
                    ind2 -= 1
                if self.ListaUb[ ind2 ] == i:
                    duplic.append( i )
                ind2 +=1
            ind += 1
        #if len(duplic) > 0:
        return duplic

    def comprobarVacios(self):  #x numero columna y posicion en fila
        x = 2; y = 1
        vacios = []
        for i in self.Lista:
            #for j in i:#enumerate(i):
            if i.modulo == None or i.modulo == '':
                vacios.append( ( Excell.getNombreColumna(y) , str(x)) )
            y += 1
            if i.estanteria == None or i.estanteria == '':
                vacios.append( ( Excell.getNombreColumna(y) , str(x)) )
            y += 1
            if i.ubicacion == None or i.ubicacion == '':
                vacios.append( ( Excell.getNombreColumna(y) , str(x)) )
            y += 1
            if i.division == None or i.division == '':
                vacios.append( ( Excell.getNombreColumna(y) , str(x)) )
            y += 1
            if i.codigo == None or i.codigo == '':
                vacios.append( ( Excell.getNombreColumna(y) , str(x)) )
            y += 1
            if i.nombre == None or i.nombre == '':
                vacios.append( ( Excell.getNombreColumna(y) , str(x)) )
            y += 1
            if i.pacto == None or i.pacto == '':
                vacios.append( ( Excell.getNombreColumna(y) , str(x)) )
            y += 1
            if i.minimo == None or i.minimo == '':
                vacios.append( ( Excell.getNombreColumna(y) , str(x)) )
            y += 1
            if i.dc == None or i.dc == '':
                vacios.append( ( Excell.getNombreColumna(y) , str(x)) )
            y += 1
            if i.gfh == None or i.gfh == '':
                vacios.append( ( Excell.getNombreColumna(y) , str(x)) )
            y += 1
            if i.dispositivo == None or i.dispositivo == '':
                vacios.append( ( Excell.getNombreColumna(y) , str(x)) )
            y += 1
            if i.hospital == None or i.hospital == '':
                vacios.append( ( Excell.getNombreColumna(y) , str(x)) )

            y = 1
            x += 1

        if len(vacios) > 0:
            vacios.append( ( str(0) , str(0)) )

        
        print('Vacios:' , vacios)
        return vacios

    def comprobarGfhDisp(self):
        l = []
        err = []
        ind = 0
        try:
            for i in self.Lista:
                l.append( ( i.gfh, i.dispositivo, i.hospital ) )
        except Exception as e:
            print('Exception en ' + str(e))
            return err
        for i in l:
            #print( 'Lista: ' + str( self.Lista[ind][9] ))
            #print( 'Tupla: ' + str( i[0] ))
            if i[0] != self.Lista[0].gfh:
                err.append( ( ind + 2 , i[0] ) )
                #print( 'Indice: ' + str( i[0] + '  ' + str( self.Lista[9] ) ) )

            if i[1] != self.Lista[0].dispositivo:
                err.append( ( ind + 2 , i[1] ) )
                #print( 'Indice: ' + str( i[1] + '  ' + str( self.Lista[ind][10] ) ) )

            if i[2] != self.Lista[0].hospital:
                err.append( ( ind + 2 , i[2] ) )
                #print( 'Indice: ' + str( i[2] + '  ' + str( self.Lista[ind][12] ) ) )
            
            ind += 1
        return err
        
    def comprobar_DC(self):
        mtx = []
        if self.Lista[1].dc == '1' or self.Lista[1].dc=='2':
            print('Configuracion Doble Cajon.')

            for i in self.Lista:
                mtx.append(i.codigo)
            aux = defaultdict(list)
            for index, item in enumerate(mtx):
                aux[item].append(index)
            result = {item: indexs for item, indexs in aux.items() if len(indexs) > 2 or len(indexs)==1}
            return result

        elif self.Lista[0].dc == 'n' or self.Lista[0].dc=='s':
            print('Configuracion Cajon Simple.')
            return mtx

        return mtx
